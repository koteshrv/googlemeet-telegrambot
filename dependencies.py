from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from rich.console import Console
from selenium import webdriver
from datetime import datetime, timedelta
from rich.table import Table
from art import *
from os import environ
import openpyxl, calendar, requests, json, time, re, discord, threading, sys, string, os, config

class color:
   PURPLE = '\033[95m'
   CYAN = '\033[96m'
   DARKCYAN = '\033[36m'
   BLUE = '\033[94m'
   GREEN = '\033[92m'
   YELLOW = '\033[93m'
   RED = '\033[91m'
   BOLD = '\033[1m'
   UNDERLINE = '\033[4m'
   END = '\033[0m'


# fetch data from json and returns data
# argument: file name of json
def fetchDataFromJSON(fileName):
	with open(config.PATH + fileName) as file:
		if file:
			data = json.load(file)
			return data
		else:
			print('Fetch failed')

# export data to json
# arguments: name of the file and data that we want to export
def sendDataToJSON(fileName, data):
	with open(config.PATH + fileName, 'w') as file:
		json.dump(data, file, indent = 4)	

data = fetchDataFromJSON('data.json')
profilePath = data['dir']['profilePath']

# rich console
# https://github.com/willmcgugan/rich
console = Console()

chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = os.environ.get("GOOGLE_CHROME_BIN")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--no-sandbox")
chrome_options.page_load_strategy = 'eager'
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_argument("--disable-extensions")
chrome_options.add_argument("--disable-popup-blocking")
chrome_options.add_argument("--user-data-dir=" + profilePath)
chrome_options.add_experimental_option("prefs", { \
"profile.default_content_setting_values.media_stream_mic": 2, # 1:allow, 2:block
"profile.default_content_setting_values.media_stream_camera": 2,
"profile.default_content_setting_values.geolocation": 2,
"profile.default_content_setting_values.notifications": 2
})

# classes and xpaths of google meet and google classroom elements
mailBoxXPath = '//*[@id="identifierId"]'
nextButtonXPath = '//*[@id="identifierNext"]'
enterPasswordBoxXPath = '//*[@id="password"]/div[1]/div/div[1]/input'
passwordNextButtonXPath = '//*[@id="passwordNext"]'
meetLinkXPath = '//*[@id="yDmH0d"]/div[4]/div[3]/div/div[1]/div/div[2]/div[2]/div/span/a'
meetLinkInCommentsXPath = '//*[@id="ow43"]/div[2]/div/div[1]/div[2]/div[1]/html-blob/span/a[1]'
dateTimeInCommentsXPath = '//*[@id="ow43"]/div[2]/div[1]/div[1]/div[1]/div[1]/span/span[1]'
classroomPostClass = 'n8F6Jd'
meetLinkClass = 'qyN25' 
warningDismissButton = '//*[@id="yDmH0d"]/div[3]/div/div[2]/div[3]/div/span/span'
membersCountBeforeJoiningClass = 'Yi3Cfd'
messageAboveJoinButtonClass = 'JMAjle'
joinButtonXPath = '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[2]/div/div[2]/div/div[1]/div[1]/span'
captionsButtonXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[2]/div/div[3]/div/span/button/span[2]'
captions = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[7]/div/div[2]/div'
membersCountXPath = '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[3]/div[2]/div/div/div[2]/div/div'

def checkStatus(var):
	logData = fetchDataFromJSON('log.json') 
	return logData["log"][str(var)]

def setStatus(var, status = True):
	logData = fetchDataFromJSON('log.json') 
	logData["log"][str(var)] = status
	sendDataToJSON('log.json', logData)

# sends message to discord
def discord(message):
	url = environ('DISCORD_WEBHOOK')
	Message = {
		"content": message
	}
	requests.post(url, data = Message)

# prints text to terminal and discord
def discordAndPrint(text):
	discord('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + text)
	print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + text)

# loading animation with text
def loadingAnimation(text = 'Loading', seconds = 5):
	animation = "|/-\\"
	idx = 0
	for i in range(seconds * 10):
	    print(animation[idx % len(animation)] + ' ' +  str(text), end="\r")
	    idx += 1
	    time.sleep(0.1)

# custom output
def printInSameLine(str1 = 'Loading', str2 = '.', sleepTime = 10, isChar = True, newLine = False, seconds = False, color = "bold green", minutes = False):
	if newLine:
		print()
		return
	sec = sleepTime
	s = ' ' * 40
	for x in range (0, int(sleepTime) + 1): 
		minutesOrSeconds = ''
		if seconds:
			minutesOrSeconds = str(sec - x)
		if minutes:
			hours = (sec - x) // 60
			minutes = sec - x - (hours * 60)
			minutesOrSeconds = str(hours) + " hr : " + str(minutes) + " min"
		b = str1 + minutesOrSeconds +  str(str2) * (x if isChar else 1)
		print(s, end = '\r')
		console.print(b, end = '\r', style = color)
		if sleepTime > 0:
			time.sleep(1)
		
		if minutes:
			time.sleep(60)
		s = ' ' * len(b) 
	if isChar:
		print('')

# returns current day with 0 index
# if current day is Monday then it returns 0, Tueday 1, Wednesday 2 ... and for Sunday 6
def findDay():
	dateAndTime = datetime.now()
	date = str(dateAndTime.day) + ' ' + str(dateAndTime.month) + ' ' + str(dateAndTime.year)
	weekDay = datetime.strptime(date, '%d %m %Y').weekday()
	return weekDay

# returns second saturday of present month
def isSecondSaturday():
	dateAndTime = datetime.now()
	month = dateAndTime.month
	year = dateAndTime.year
	day = dateAndTime.day
	monthCalendar = calendar.monthcalendar(year, month)
	if monthCalendar[0][calendar.SATURDAY]:
		secondSaturday = monthCalendar[1][calendar.SATURDAY]
	else:
		secondSaturday = monthCalendar[2][calendar.SATURDAY]
	return secondSaturday

# updates the holiday list
# when no key and value are given, then it checks for second saturday and sunday
# when key and value are given then it updates the holiday list with key and value and also check for second saturday and sunday
# when remove is true, then it removes the key and value from the holidays list
def updateholidaysList(key = None, value = None, remove = False):
	jsonData = fetchDataFromJSON('log.json')
	holidaysDict = jsonData["holidaysList"]
	if remove:
		del holidaysDict[key]
		console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Deleted '+ key + ' from holidays list successfully', style = "gold3")
	else :
		dateAndTime = datetime.now()
		day = dateAndTime.day
		secondSaturday = isSecondSaturday()
		if secondSaturday >= day:
			holidaysDict[str(secondSaturday)] = 'Second Saturday'
		if findDay() == 6:
			holidaysDict[str(day)] = 'Sunday'
		if not (key == None and value == None):
			holidaysDict[str(key)] = value
		for holidayDate in list(holidaysDict):
			if int(holidayDate) < day:
				holidaysDict.pop(holidayDate)

	jsonData["holidaysList"].update(holidaysDict)
	sendDataToJSON('log.json', jsonData)

# loads the time table from the excel sheet and then updates the classes today
def loadTimeTable():
	classTimeTableLocation = data['dir']['classTimeTableLocation']
	timetablewb = openpyxl.load_workbook(classTimeTableLocation) 
	sheet = timetablewb.active
	maxColumn = sheet.max_column
	maxRow = sheet.max_row
	completeTimeTable = {}
	for i in range(1, maxRow + 1):
		rowValues = []
		keyValue = sheet.cell(row = i, column = 1)
		for j in range(2, maxColumn + 1):
			cellValue = sheet.cell(row = i, column = j)
			if cellValue.value != None:
				rowValues.append(cellValue.value)
		completeTimeTable[keyValue.value] = rowValues
	jsonData = fetchDataFromJSON('log.json')
	jsonData["completeTimeTable"].update(completeTimeTable)
	sendDataToJSON('log.json', jsonData)
	classesToday(printHoliday = False)

# returns the status of the classwork
# if no class is going on, then it returns False which means that classes are scheduled later
# if classwork is completed then it returns True which means all classes are already completed
# if classwork is going on, then it returns -1
def classStatus():
	jsonData = fetchDataFromJSON('log.json')
	todaysTimeTable = jsonData["todaysTimeTable"]
	timings = list(todaysTimeTable.keys())
	startTime = timings[0][:5]
	endTime = timings[-1][8:]
	h, m, s = str(datetime.now().time()).split(':')
	presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
	classstartTime = timedelta(hours = int(startTime[:2]), minutes = int(startTime[3:5]))
	classendTime = timedelta(hours = int(endTime[:2]), minutes = int(endTime[3:]))
	if presentTime < classstartTime:
		return False
	if presentTime > classstartTime and presentTime < classendTime:	
		return -1
	if not presentTime < classendTime:
		return True

# makes a schedule  for classes today and prints  
def classesToday(printTable = False, printHoliday = True):
	date = findDay()
	classes = []
	updateholidaysList()
	jsonData = fetchDataFromJSON('log.json')
	holidays = jsonData["holidaysList"]
	dateAndTime = datetime.now()
	day = dateAndTime.day
	weekDay = dateAndTime.today().strftime('%A')
	if str(day) in holidays and printHoliday:
		console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + "\n\tToday is a holiday due to ", style = "bold cyan", end = '')
		console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + holidays[str(day)] + '\n', style = "bold yellow")
	else:
		classes = []
		classtime = []
		classesToday = jsonData["completeTimeTable"][weekDay]
		timings = jsonData["completeTimeTable"]["Timings"]
		timings = timings[:len(classesToday)]
		prevClass = ''
		for i in range(len(classesToday)):
			if classesToday[i] == prevClass:
				classtime[-1] = classtime[-1][:8] + timings[i][8:]
			else :
				classes.append(classesToday[i])
				classtime.append(timings[i])
				prevClass = classesToday[i]
		t = {classtime[i]: classes[i] for i in range(len(classtime))}
		jsonData = fetchDataFromJSON('log.json')
		classLoginLog = jsonData["log"]["classStatus"]
		classesAlreadyAttended = list(classLoginLog.keys())
		jsonData["todaysTimeTable"] = t
		sendDataToJSON('log.json', jsonData)
		if printTable:
			table = Table(show_lines=True)
			table.add_column("Timings", justify = "center", style = "cyan", no_wrap = True)
			table.add_column("Class", justify = "center", style = "green")
			table.add_column("Status", justify = "center", style = "magenta")
			classFlag = False
			h, m, s = str(datetime.now().time()).split(':')
			presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
			for i in range(len(classtime)):
				startTime = timedelta(hours = int(classtime[i][0:2]), minutes = int(classtime[i][3:5]))
				endTime = timedelta(hours = int(classtime[i][8:10]), minutes = int(classtime[i][11:]))
				if (not classFlag) and (presentTime < startTime):
					table.add_row(classtime[i], classes[i], "[bold magenta] Scheduled [green]:hourglass:")
				elif (presentTime < endTime and presentTime > startTime) and (not(classtime[i] in classesAlreadyAttended)):
					table.add_row(classtime[i], classes[i], "[bold magenta] ONGOING [green]:hourglass:")
					classFlag = True
				else:
					if classFlag == True:
						table.add_row(classtime[i], classes[i])
					else :
						table.add_row(classtime[i], classes[i], "[bold magenta]COMPLETED [green]:heavy_check_mark:")
			console.print(table)
			print()

# returns current class 
# if no class is going on or if the current class is already attended then it will return None
def whichClass(nextClass = False, nextClassTime = False, currentClassTime = False) :
	loadTimeTable()
	jsonData = fetchDataFromJSON('log.json')
	subjects = jsonData["todaysTimeTable"]
	classLoginLog = jsonData["log"]["classStatus"]
	classesAlreadyAttended = list(classLoginLog.keys())
	nextClassFlag = False
	h, m, s = str(datetime.now().time()).split(':')
	presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
	for i in subjects:
		startTime = timedelta(hours = int(i[0:2]), minutes = int(i[3:5]))
		endTime = timedelta(hours = int(i[8:10]), minutes = int(i[11:]))
		if nextClassFlag and (not nextClassTime):
			return subjects[i]
		if nextClassFlag and nextClassTime:
			return (i, subjects[i])
		if nextClass:
			if presentTime < endTime and presentTime > startTime:
				nextClassFlag = True
				continue
		elif  presentTime < endTime and presentTime > startTime and (not(i in classesAlreadyAttended)):	
			if currentClassTime:
				return (i, subjects[i])
			return subjects[i]	
	if nextClassTime or currentClassTime:
		return (None, None)
	return None

# prints complete time table
def displayTimeTable():
	loadTimeTable()
	jsonData = fetchDataFromJSON('log.json')
	t = jsonData["completeTimeTable"]
	dateAndTime = datetime.now()
	weekDay = dateAndTime.today().strftime('%A')
	table = Table(show_lines=True)
	timetableKeys = list(t.keys())
	table.add_column(timetableKeys[0], justify = "center", style = "cyan", no_wrap = True)
	for i in range(len(t[timetableKeys[0]])):
		table.add_column(t[timetableKeys[0]][i], justify = "center", style = "yellow", no_wrap = True)
	for i in timetableKeys[1:]:
		tempList = [i] + t[i]
		if weekDay == i:
			table.add_row(*tempList, style = "bold")
		else:
			table.add_row(*tempList)
	console.print(table)

# displays holidays list
def displayHolidaysList():
	updateholidaysList()
	data = fetchDataFromJSON('log.json')
	holidayListKeys = list(data["holidaysList"].keys())
	table = Table(title = "[blink2 bold dodger_blue1]Holidays List", show_lines=True)
	table.add_column("[dark_orange]Date", justify = "center", style = "yellow3", no_wrap = True)
	table.add_column("[dark_orange]Occasion", justify = "center", style = "yellow3", no_wrap = True)
	holidaysFlag = False
	for key in holidayListKeys:
		value = data["holidaysList"][key]
		l = [key, value]
		table.add_row(*l)
		holidaysFlag = True
	if holidaysFlag:
		console.print(table)
	else :
		console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + "You dont have any holidays", style = "bold gold1")

# reverts to the original timetable
def revertTimeTable():
	jsonData = fetchDataFromJSON('log.json')
	totalUpdates = len(list(jsonData["tempTimetableUpdate"].keys()))
	if totalUpdates > 0:
		for i in range(totalUpdates):
			day = jsonData["tempTimetableUpdate"][str(i)]["day"]
			period = jsonData["tempTimetableUpdate"][str(i)]["period"]
			classToUpdate = jsonData["tempTimetableUpdate"][str(i)]["previousClass"]
			updateTimeTable(day, period, classToUpdate)
		jsonData["tempTimetableUpdate"] = {}
		sendDataToJSON('log.json', jsonData)

# updates the time table by taking day, period and class to update
# day should be string of any format. eg: "Monday", "monday", "MONday"
# period should be passed as stored in excel sheet
# class that you want to replace with
# if you want to remove a class then just pass the class with ''
def updateTimeTable(day, period, classToUpdate = None, previousClass = False):
	classTimeTableLocation = data['dir']['classTimeTableLocation']
	timetablewb = openpyxl.load_workbook(classTimeTableLocation) 
	sheet = timetablewb.active
	day = day.lower()
	log = fetchDataFromJSON('log.json')
	i, val, asciiVal = 1, 2, 66
	rowValues, colValues, colDict = {}, {}, {}
	for key in list(log['completeTimeTable'].keys())[1:] :
		key = key.lower()
		rowValues[key] = str(val)
		colValues[key] = chr(asciiVal)
		colDict[str(i)] = key
		i += 1
		asciiVal += 1	
		val += 1	
	
	prevClass =  sheet[colValues[colDict[period]] + str(rowValues[day])].value
	sheet[colValues[colDict[period]] + str(rowValues[day])] = classToUpdate
	timetablewb.save(classTimeTableLocation)
	if previousClass:
		return prevClass
	
# prints log data
def printLog(discordServer = False):
	logData = fetchDataFromJSON('log.json')
	if discordServer:
		return json.dumps(logData, indent = 4)
	print(json.dumps(logData, indent = 4))

# prints the arguments and their uses 
def helpFunction():
	table = Table(show_lines=True)
	table.add_column("Arguments", justify = "center", style = "yellow3", no_wrap = True)
	table.add_column("Details", justify = "center", style = "yellow3")
	table.add_row("no arguments", "runs the main program")
	table.add_row("--t", "displays todays timetable")
	table.add_row("--h", "displays holidays list")
	table.add_row("--c", "displays present class")
	table.add_row("--l", "uses class url to join the class immediately")
	table.add_row("--log", "displays log")
	table.add_row("--l 'time'", "uses class url to join the class and schedules at specified 'time'")
	table.add_row("--h -a", "add new holiday to the list and prints")
	table.add_row("--h -r", "removes holiday from the list and prints")
	table.add_row("--t -f", "displays complete timetable fetched from excel sheet")
	table.add_row("--t -u", "changes timetable and displays complete timetable fetched from excel sheet")
	table.add_row("--t -ut", "changes timetable temporarily and updates back the original timetable next time")
	console.print(table)

# returns the count of members joined before joining the class
def membersAlreadyJoinedCount(text):
	if text == 'No one else is here':
		return 0
	count = 0
	numberFromText = [int(i) for i in text.split() if i.isdigit()]
	commasCount = text.count(',')
	andCount = text.count('and')
	if len(numberFromText) > 0:
		count = numberFromText[0]
		andCount = 0
	count = count + commasCount + andCount + 1
	return count

# joins the class of given subject
def joinClass(subject = None, URL = None, loginTime = None):
	# used when we want to implicitly wait 
	# time = 90/60 = 1.5 minutes
	loadDriver()
	wait = WebDriverWait(driver, 90)
	if URL == None:
		log = {}
		subject = subject.upper()
		url = data['classroomLinks'][subject]
		discordAndPrint('Opening ' + subject + ' classroom in new tab')
		while True:
			try:
				driver.get(url)
				break
			except Exception:
				discordAndPrint('Timeout exception occured! So closing current driver and trying again')
				driver.close()
				driver.quit()
				loadDriver()
		loadingAnimation()
		discordAndPrint('Waiting for Google Meet link for ' + subject + ' class')
		linkPostedSeperatelyInAnnouncementTab = data['otherData']['linkPostedSeperatelyInAnnouncementTab']
		# if subject link is posted seperately in announcement tab
		if subject in linkPostedSeperatelyInAnnouncementTab:	
			discordAndPrint(subject + ' class Link is posted in announcement tab')
			discordAndPrint('So trying to fetch data from announcement tab')
			previousPostData = None
			while True:
				if abortBotIfTriggered():
					return
				# from the below fetched data
				# check the date is matching before joining
				# if the link is posted today, then the element stores the time for eg: "12.06 AM"
				# if the link is not posted today, then element stores the day for eg: "May 11"
				while True:
					try:
						announcementTabData = str(driver.find_element_by_class_name(classroomPostClass).text)
						announcementTabpostedDateTime = str(driver.find_element_by_xpath(dateTimeInCommentsXPath).text)
						break
					except Exception:
						discordAndPrint('Exception occured when trying to fetch data. So trying again in 30 seconds')
						for i in range(30):
							if abortBotIfTriggered():
								return
							time.sleep(1)
				if previousPostData != announcementTabData :
					print('\n' + '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + color.BOLD +'Fetched Data' + color.END + '\n')
					print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + color.BOLD + color.YELLOW + announcementTabData + color.END + '\n')
					discord('__**Fetched Data**__')
					discord('```' + announcementTabData + '```')
				previousPostData = announcementTabData
				# fetching url from annoucement tab data
				# until url is fetched, the page loads for every 10 seconds
				classURL = re.search("(?P<url>https?://[^\s]+)", announcementTabData).group("url")
				if not announcementTabpostedDateTime[8].isalpha():
					if (classURL[:24] == 'https://meet.google.com/') :
						discordAndPrint('Fetched class link from the google classroom')
						discordAndPrint('Opening ' + classURL)
						while True:
							try:
								driver.get(classURL)
								break
							except Exception:
								discordAndPrint('Timeout exception occured! So closing current driver and trying again')
								driver.close()
								driver.quit()
								loadDriver()
						loadingAnimation()
						break
					else:
						discordAndPrint('Fetching link failed or link not posted')
				else :
					driver.refresh()
					discordAndPrint('Waiting for Todays link. Trying again in 10 seconds')
					for i in range(10):
						time.sleep(1)
						if abortBotIfTriggered():
							return
		# fetches data from link posted in google classroom
		# if link is not available and it loads and check for every 10 seconds
		else :
			while True:
				try:
					if abortBotIfTriggered():
						return
					classData = driver.find_element_by_class_name(meetLinkClass).text
					classURL = re.search("(?P<url>https?://[^\s]+)", classData).group("url")
					if classURL[:24] == 'https://meet.google.com/':
						discordAndPrint('Fetched class link from the google classroom')
						discordAndPrint('Opening ' + classURL)
						while True:
							try:
								driver.get(classURL)
								break
							except Exception:
								discordAndPrint('Timeout exception occured! So closing current driver and trying again')
								driver.close()
								driver.quit()
								loadDriver()
						discordAndPrint('Opened meet link')
						loadingAnimation()
						break
				except AttributeError:
					if abortBotIfTriggered():
						return
					discordAndPrint('Meet link not available to fetch. Trying again in 10 seconds')
					driver.refresh()
					for i in range(10):
						time.sleep(1)
						if abortBotIfTriggered():
							return

	else:
		if loginTime == None:
			print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Opening ', URL)
			while True:
				try:
					driver.get(URL)
					break
				except Exception:
					discordAndPrint('Timeout exception occured! So closing current driver and trying again')
					driver.close()
					driver.quit()
					loadDriver()
			loadingAnimation()
		else:
			h, m, s = str(datetime.now().time()).split(':')
			presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
			timeLeftForNextClass = (timedelta(hours = loginTime[:2], minutes = loginTime[-2:]) - presentTime).total_seconds()
			discordAndPrint('Class is scheduled successfully. Will join the class in ' + str(timedelta(seconds = timeLeftForNextClass)))
			for i in range(timeLeftForNextClass):
				time.sleep(1)
				if abortBotIfTriggered():
					return
			
		discordAndPrint('Opening ' + URL)
		while True:
			try:
				driver.get(URL)
				break
			except Exception:
				discordAndPrint('Timeout exception occured! So closing current driver and trying again')
				driver.close()
				driver.quit()
				loadDriver()
		discordAndPrint('Opened meet link')
		loadingAnimation()


	discordAndPrint('Pressing dismiss button')
	warningDismiss = driver.find_element_by_xpath(warningDismissButton).click()
	time.sleep(3)
	# fetching count of members already joined
	# if members count is not available then is sets the  minCountToJoinConsidered to False
	try :
		membersCountBeforeJoiningData = driver.find_element_by_class_name(membersCountBeforeJoiningClass).text
		print('\n' + '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + color.BOLD + color.YELLOW + str(membersCountBeforeJoiningData) + color.END + '\n')
		discord(str(membersCountBeforeJoiningData))
		joinedMembers = membersAlreadyJoinedCount(membersCountBeforeJoiningData)
		minCountToJoinConsidered = True
		minCountToJoin = data['otherData']['minCountToJoin']
	except NoSuchElementException:
		#messageAboveJoinButtonData = driver.find_element_by_class_name(messageAboveJoinButtonClass).text
		discordAndPrint("Members count not available! So joining the class without considering 'minCountToJoin'")
		minCountToJoinConsidered = False
	# checking for minCountToJoin to join the class
	while True:
		if abortBotIfTriggered():
			return
		if not minCountToJoinConsidered:
			break
		if joinedMembers >= minCountToJoin: 
			discordAndPrint('More than ' + str(minCountToJoin) + ' members already joined')
			discordAndPrint('Joining the class now')
			break
		else :
			if joinedMembers == 0:
				discordAndPrint('No one joined. Will try for every 10 seconds')
				for i in range(10):
					time.sleep(1)
					if abortBotIfTriggered():
						return
			else :
				discordAndPrint('Only ' + str(joinedMembers) + ' joined')
				discordAndPrint('Waiting for ' + str(minCountToJoin - joinedMembers) + ' more students to join the class')
				discord('Trying again in 10 seconds')
				for i in range(10):
					time.sleep(1)
					if abortBotIfTriggered():
						return
				membersCountBeforeJoiningData = driver.find_element_by_class_name(membersCountBeforeJoiningClass).text
				joinedMembers = membersAlreadyJoinedCount(membersCountBeforeJoiningData)
	# waits until join button is clickable
	element = wait.until(EC.element_to_be_clickable((By.XPATH, joinButtonXPath)))
	element.click()
	discordAndPrint('Pressing join button')
	# waits until caption button is clickable and turn on captions
	element = wait.until(EC.element_to_be_clickable((By.XPATH, captionsButtonXPath)))
	element.click()
	discordAndPrint('Turning on captions')
	alertSound(False)
	if URL == None:
		# sending class joining time to discord
		discord("Joined " + subject + " class at " + str(datetime.now().time())[:8])
		classTime = whichClass(currentClassTime = True)[0]
		joiningLeavingTimeDict = {}
		joiningLeavingTimeDict["joining time"] = str(datetime.now().time())
		if classTime in log:
			log[classTime].update(joiningLeavingTimeDict)
		else :
			log[classTime] = joiningLeavingTimeDict
		logData = fetchDataFromJSON('log.json')
		logData["log"]["joiningLeavingTime"].update(log)
		sendDataToJSON('log.json', logData)
		time.sleep(3)

	else :
		# sending class joining time to discord
		discord('Joined the class with ' + URL + ' successfully at ' + str(datetime.now().time())[:8])
	# counting number of students joined 
	count = driver.find_element_by_xpath(membersCountXPath).text
	flag = False
	minCountToLeave = data['otherData']['minCountToLeave']
	alertWords = data['otherData']['alertWords']
	autoReply = data['otherData']['autoReply']
	logData = fetchDataFromJSON('log.json')
	# Reads the text from captions until str(count) > minCountToLeave:
	while True:
		messageFromDiscord = fetchDataFromJSON('log.json')
		messageFromDiscord = messageFromDiscord["log"]["messageToSendFromDiscordToMeet"]
		if abortBotIfTriggered():
			return
		try:
			count = driver.find_element_by_xpath(membersCountXPath).text
			updateMembersCount(count)
		finally:
			pass
		usedPrintInSameLine = False
		if not checkStatus('discordServer'):
			printInSameLine('Members Count: ', count, sleepTime = 0, isChar = False)
			usedPrintInSameLine = True
		
		try:
			# flag is used to check when class count reaches above minCountToLeave
			# when it is set to true it implies that it is waiting to leave the class when count reaches below minCountToLeave
			if count > str(minCountToLeave):
				flag = True
			elems = driver.find_element_by_xpath(captions)
			captionTextLower = str(elems.text).translate(str.maketrans('', '', string.punctuation)).lower()
			# if alert word is found in captions then it plays an alert sound for soundFrequency times and then sends alert message to discord
			# if you want to enable auto replay when this triggers then uncomment the line below. This sends the response message that is given in data.json file
			for word in alertWords:
				if word in captionTextLower:
					if usedPrintInSameLine:
						printInSameLine(newLine = True)
					print(text2art("ALERT", font = "small")) 
					discordAndPrint("ALERT! Some one called you at " + str(datetime.now().time())[:8])
					discordAndPrint(captionTextLower)
					discordAndPrint("Triggered word: " + word)
					alertSound() # alert sound for soundCount times
					for i in range(10):
						if abortBotIfTriggered():
							return 
						time.sleep(1)	

			# leaves the class when class count is less than minCountToLeave
			if (count < str(minCountToLeave) and flag) or data["otherData"]["leaveAtTime"]:
				if (count < str(minCountToLeave) and flag) and usedPrintInSameLine:
					printInSameLine(newLine = True)
				if data["otherData"]["leaveAtTime"]:
					waitUntilClassEnds(whichClass(currentClassTime = True))
				alertSound(frequency = False)
				if URL == None:
					discord("Left the " + subject + " class at " + str(datetime.now().time())[:8])
					joiningLeavingTimeDict["leaving time"] = str(datetime.now().time())
					log[classTime].update(joiningLeavingTimeDict)
					logData = fetchDataFromJSON('log.json')
					logData["log"]["joiningLeavingTime"].update(log)
					sendDataToJSON('log.json', logData)
					driver.close()
					updateMembersCount(None)
					loadingAnimation('Left '+ subject + ' class')
					console.print('\n' + '[' + str(datetime.now().strftime("%H:%M:%S")) + ']  Left the class successfully', style = "bold red", end = '\r')	
				else :
					discord('Left the class of url ' + URL + ' successfully at ' + str(datetime.now().time())[:8])
					driver.close()
					updateMembersCount(None)
					console.print('\n'+ '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' +'Left the class successfully', style = "bold red")	
				setStatus("status", False)
				setStatus("membersCount", None)
				break
				
					
		except (NoSuchElementException, StaleElementReferenceException):
			# flag is used to check when class count reaches above minCountToLeave
			# when it is set to true it implies that it is waiting to leave the class when count reaches below minCountToLeave
			if count > str(minCountToLeave):
				flag = True

			if (count < str(minCountToLeave) and flag) or data["otherData"]["leaveAtTime"]:
				if (count < str(minCountToLeave) and flag) and usedPrintInSameLine:
					printInSameLine(newLine = True)
				if data["otherData"]["leaveAtTime"]:
					waitUntilClassEnds(whichClass(currentClassTime = True))
				alertSound(frequency = False)
				if URL == None:
					discord("Left the " + subject + " class at " + str(datetime.now().time())[:8])
					joiningLeavingTimeDict["leaving time"] = str(datetime.now().time())
					log[classTime].update(joiningLeavingTimeDict)
					logData = fetchDataFromJSON('log.json')
					logData["log"]["joiningLeavingTime"].update(log)
					sendDataToJSON('log.json', logData)
					driver.close()
					updateMembersCount(None)
					loadingAnimation('Left '+ subject + ' class')
					console.print('\n' + '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Left the class successfully', style = "bold red", end = '\r')	
				else :
					discord('Left the class of url ' + URL + ' successfully at ' + str(datetime.now().time())[:8])
					driver.close()
					updateMembersCount(None)
					console.print('\n' + '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Left the class successfully', style = "bold red")	
				setStatus("status", False)
				setStatus("membersCount", None)
				break				

	
# checks for abort call from discord server
def abortBotIfTriggered():
	if checkStatus("stop") and checkStatus("abort"):
		return 1
	if checkStatus("stop"):
		print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Abort called from discord server.')
		print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'So force stopping the bot')
		updateMembersCount(None)
		discord('Force stopped the meet successfully')
		try:
			driver.quit()
		except Exception:
			pass
		setStatus("status", False)
		setStatus("membersCount", None)
		setStatus("abort")
		return 1
	return 0

# plays alert sound for soundFrequency times where soundFrequency is stored in data.json
def alertSound(frequency = True):
	if frequency:
		soundFrequency = data['otherData']['soundFrequency']
		for i in range(soundFrequency):
			#winsound.Beep(500, 1000)
			time.sleep(1)	
	else:
		#winsound.Beep(500, 1000)
		pass

# launches the chrome driver 
def loadDriver():
	global driver
	pathToChromeDriver = data['dir']['pathToChromeDriver']
	driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=chrome_options)
	driver.maximize_window()
	driver.set_page_load_timeout(10)
	discordAndPrint('Driver loaded successfully!')
	setStatus("abort", False)

# used to login to gmail 
# not used here as we used chrome profile
def login():
	discordAndPrint('Logging into ' + color.BOLD + 'Google account' + color.END)
	while True:
		try:
			driver.get('https://classroom.google.com/?emr=0')
			break
		except Exception:
			discordAndPrint('Timeout exception occured! So closing current driver and trying again')
			driver.close()
			driver.quit()
			loadDriver()
	time.sleep(3)
	mailAddress = data['credentials']['mailAddress']
	password = data['credentials']['password']
	print('Entering mail address')
	mailBox = driver.find_element_by_xpath(mailBoxXPath)
	driver.implicitly_wait(10)
	mailBox.send_keys(mailAddress)
	driver.find_element_by_xpath(nextButtonXPath).click()
	driver.implicitly_wait(10)
	time.sleep(2)
	print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + color.BOLD + 'Entering password' + color.END)
	passwordBox = driver.find_element_by_xpath(enterPasswordBoxXPath)
	driver.implicitly_wait(10)
	passwordBox.send_keys(password)
	driver.find_element_by_xpath(passwordNextButtonXPath).click()
	driver.implicitly_wait(10)
	time.sleep(2)
	loadingAnimation(seconds = 10)
	discordAndPrint('Login Successful')

# googlemeetbot main method thread
def googlemeetbotThread():
	setStatus("stop", False)
	setStatus("abort", False)
	setStatus("status")
	botThread = threading.Thread(target = googlemeetbotFunction)
	botThread.start()

# join class method thread
def joinClassThread(classLink, schedule = None):
	setStatus("stop", False)
	setStatus("abort", False)
	setStatus("status")
	if schedule == None:
		joinClassThread = threading.Thread(target = joinClass, args=(None, classLink, None, ))
	else :
		joinClassThread = threading.Thread(target = joinClass, args=(None, classLink, schedule, ))
	joinClassThread.start()
	
# updates member count to log.json
def updateMembersCount(count):
	logData = fetchDataFromJSON('log.json')
	logData["log"]["membersCount"] = count
	sendDataToJSON('log.json', logData)

def waitUntilClassEnds(presentClass):
	start_h, start_m = map(int, presentClass[0][8:].split(':'))
	h, m, s = str(datetime.now().time()).split(':')
	endTime = timedelta(hours = start_h, minutes = start_m)
	currentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
	if endTime > currentTime:
		timeLeftForNextClass = (endTime - currentTime).total_seconds()
		discordAndPrint('Time left : ' + str(timedelta(seconds = timeLeftForNextClass)))
		while True:
			try:
				h, m, s = str(datetime.now().time()).split(':')
				currentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
				if currentTime > endTime:
					break
				messageFromDiscord = fetchDataFromJSON('log.json')
				messageFromDiscord = messageFromDiscord["log"]["messageToSendFromDiscordToMeet"]
				if abortBotIfTriggered():
					break
			except Exception as e:
				discordAndPrint(str(e))
		if abortBotIfTriggered():
			return
		

# googlemeetbot function	
def googlemeetbotFunction():
	print(text2art("googlemeetbot", font = "small"))
	# Checking for previous day log
	# If previous day log is present then it's cleared
	dateAndTime = datetime.now()
	day = dateAndTime.day
	weekDay = dateAndTime.today().strftime('%A')
	jsonData = fetchDataFromJSON('log.json')
	if jsonData["log"]["day"] != weekDay:
		jsonData["log"]["day"] = weekDay
		jsonData["log"]["joiningLeavingTime"] = {}
		jsonData["log"]["classStatus"] = {}
	sendDataToJSON('log.json', jsonData)

	# checking for holiday
	dateAndTime = datetime.now()
	day = dateAndTime.day
	updateholidaysList()
	loadTimeTable()
	jsonData = fetchDataFromJSON('log.json')
	holidaysDict = jsonData["holidaysList"]
	classesTodayData = jsonData["todaysTimeTable"]
	for holidayDate in holidaysDict:
		if str(day) == holidayDate:
			console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'No classes today due to '+ holidaysDict[holidayDate], style = "bold red")
			console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + '[bold][green]' + 'Done\n')
			return

	# todays class list from the present time
	# for suppose if we have classes from 9:00 to 4:00 and if we run this script at 10:00 
	# then it will consider classes from 10:00 into classList
	classesList = []
	flag = False
	for timings in classesTodayData:
		# if current time is less than class start time then we should add all periods to the list
		h, m, s = str(datetime.now().time()).split(':')
		presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
		startTime = timedelta(hours = int(timings[0:2]), minutes = int(timings[3:5]))
		endTime = timedelta(hours = int(timings[8:10]), minutes = int(timings[11:]))
		
		if presentTime < startTime: 
			flag = True
		# if we have a class in current time then we should add classes from now to the list
		if presentTime > startTime and presentTime < endTime:
			flag = True
		if flag:
			classesList.append(timings + ' '+ classesTodayData[timings])

	completedClassesCount = len(classesTodayData) - len(classesList)

	classesToday(printTable = True, printHoliday = False)

	# get status for class joining
	# if status is True: classwork is completed
	# if status is false: classwork is scheduled for today
	# if status if -1: classwork is going on at the moment
	status = classStatus()

	# if status is true: classwork is completed for today
	if status == True:
		console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'All classes attended for today', style = "blink2 bold red")
		discord('All classes attended for today')
		return

	# if status will be false when current time is less than start time of college so we wait here until status becomes -1	
	if status == False:
		jsonData = fetchDataFromJSON('log.json')
		todaysTimeTable = jsonData["todaysTimeTable"]
		timings = list(todaysTimeTable.keys())
		start_h, start_m = map(int, timings[0][:5].split(':'))
		h, m, s = str(datetime.now().time()).split(':')
		timeToSleep = (timedelta(hours = start_h, minutes = start_m) - timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))).total_seconds()
		discordAndPrint('You are early for the class. So I am sleeping for the next ' + str(timedelta(seconds = timeToSleep)))
		# sleeps for the next timeToSleep minutes and if killbot triggered from discord server then it aborts
		for i in range(int(timeToSleep)):
			time.sleep(1)
			if abortBotIfTriggered():
				return
		status = classStatus()

	# if today is not a holiday and status is -1 that is classwork is going on 
	if status == -1 or status == False:
		classNow = whichClass()
		totalclassesTodayData = len(classesList)

		# joins all classes that are stored in classesList
		for i in range(totalclassesTodayData):
			classNow = whichClass()

			# if current class is "None" then it will wait until next class
			# current class returns "None" when we dont have any period or if we already attended the class
			if classNow == None:
				nextClass = whichClass(nextClass = True, nextClassTime = True)
				start_h, start_m = map(int, nextClass[0][:5].split(':'))
				h, m, s = str(datetime.now().time()).split(':')
				timeLeftForNextClass = (timedelta(hours = start_h, minutes = start_m) - timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))).total_seconds()
				discordAndPrint('No class at the moment. Will try again in ' +  str(timedelta(seconds = timeLeftForNextClass)))
				for i in range(int(timeLeftForNextClass)):
					time.sleep(1)
					if abortBotIfTriggered():
						return

			# if current class returns "Lunch" then it sleeps until next class
			if classNow == "Lunch":
				nextClass = whichClass(nextClass = True, nextClassTime = True)
				start_h, start_m = map(int, nextClass[0][:5].split(':'))
				h, m, s = str(datetime.now().time()).split(':')
				timeLeftForNextClass = (timedelta(hours = start_h, minutes = start_m) - timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))).total_seconds()		
				discordAndPrint('Lunch Time. Will join next class in ' + str(timedelta(seconds = timeLeftForNextClass)))
				for i in range(int(timeLeftForNextClass)):
					time.sleep(1)
					if abortBotIfTriggered():
						return

			# joins current class and updates the log
			classNow = whichClass()	
			while classNow == None or classNow == "Lunch":
				classNow = whichClass()
				if abortBotIfTriggered():
					return
			console.print('\n' + '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + classNow, style = "bold bright_yellow", end = '')
			console.print(' class is going on at the moment\n', style = "bold")
			print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Trying to join ' + classNow + ' class')
			joinClass(subject = classNow)
			if abortBotIfTriggered():
				return
			jsonData = fetchDataFromJSON('log.json')
			t = jsonData["todaysTimeTable"]
			classTime = list(t.keys())
			periods = {}
			if classTime[i + completedClassesCount] in periods:
				periods[classTime[i + completedClassesCount]].update(classNow)
			else :
				periods[classTime[i + completedClassesCount]] = classNow

			jsonData["log"]["classStatus"].update(periods)
			sendDataToJSON('log.json', jsonData)

		# reverting the classes if previously changed
		# if the timetable is temporarily changed then its reverted to original 
		revertTimeTable()

		console.print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + "Attened all classes successfully!", style = "BLINK2 bold red")

# used to login to gmail 
# we use this only once while setting chrome profile
def login(mailAddress, password):
	pathToChromeDriver = data['dir']['pathToChromeDriver']
	driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=chrome_options)
	driver.maximize_window()
	print('Logging into ' + color.BOLD + 'Google account' + color.END)
	driver.get('https://accounts.google.com/servicelogin')
	print('Entering mail address')
	mailBox = driver.find_element_by_xpath(mailBoxXPath)
	driver.implicitly_wait(10)
	mailBox.send_keys(mailAddress)
	driver.save_screenshot("image.png")
	driver.find_element_by_xpath(nextButtonXPath).click()
	driver.implicitly_wait(10)
	print(color.BOLD + 'Entering password' + color.END)
	passwordBox = driver.find_element_by_xpath(enterPasswordBoxXPath)
	driver.implicitly_wait(10)
	passwordBox.send_keys(password)
	driver.find_element_by_xpath(passwordNextButtonXPath).click()
	driver.implicitly_wait(10)
	print('Login Successful')
	return driver
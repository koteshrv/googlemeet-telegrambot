from copy import Error
from datetime import datetime, timedelta
from dependencies import driver
from telegram import ChatAction
from art import *
from os import execl
from sys import executable
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import openpyxl, calendar, requests, json, time, re, sys, string, os, pickle, config

# fetch data from json and returns data
# argument: file name of json
def fetchDataFromJSON(fileName):
	with open(config.DEFAULT_LOCATION + fileName) as file:
		if file:
			data = json.load(file)
			return data
		else:
			print('Fetch failed')

# export data to json
# arguments: name of the file and data that we want to export
def sendDataToJSON(fileName, data):
	with open(config.DEFAULT_LOCATION + fileName, 'w') as file:
		json.dump(data, file, indent = 4)	

data = fetchDataFromJSON('data.json')

def checkStatus(var):
	logData = fetchDataFromJSON('log.json') 
	return logData["log"][str(var)]

def setStatus(var, status = True):
	logData = fetchDataFromJSON('log.json') 
	logData["log"][str(var)] = status
	sendDataToJSON('log.json', logData)

# sends message to discord
def sendToDiscord(message):
	webhook = config.DISCORD_WEBHOOK
	Message = {
		"content": '[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + message
	}
	requests.post(webhook, data = Message)

# prints text to terminal and discord
def discordAndPrint(text, log = True):
	sendToDiscord(text)
	print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + text)

# returns current day with 0 index
# if current day is Monday then it returns 0, Tueday 1, Wednesday 2 ... and for Sunday 6
def findDay():
	dateAndTime = datetime.now()
	date = str(dateAndTime.day) + ' ' + str(dateAndTime.month) + ' ' + str(dateAndTime.year)
	weekDay = datetime.strptime(date, '%d %m %Y').weekday()
	return weekDay

# returns second saturday of present month
def isSecondSaturday():
	dateAndTime = datetime.now()
	month = dateAndTime.month
	year = dateAndTime.year
	day = dateAndTime.day
	monthCalendar = calendar.monthcalendar(year, month)
	if monthCalendar[0][calendar.SATURDAY]:
		secondSaturday = monthCalendar[1][calendar.SATURDAY]
	else:
		secondSaturday = monthCalendar[2][calendar.SATURDAY]
	return secondSaturday

# updates the holiday list
# when no key and value are given, then it checks for second saturday and sunday
# when key and value are given then it updates the holiday list with key and value and also check for second saturday and sunday
# when remove is true, then it removes the key and value from the holidays list
def updateholidaysList(key = None, value = None, remove = False):
	jsonData = fetchDataFromJSON('log.json')
	holidaysDict = jsonData["holidaysList"]
	if remove:
		del holidaysDict[key]
		print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + 'Deleted '+ key + ' from holidays list successfully')
	else :
		dateAndTime = datetime.now()
		day = dateAndTime.day
		secondSaturday = isSecondSaturday()
		if secondSaturday >= day:
			holidaysDict[str(secondSaturday)] = 'Second Saturday'
		if findDay() == 6:
			holidaysDict[str(day)] = 'Sunday'
		if not (key == None and value == None):
			holidaysDict[str(key)] = value
		for holidayDate in list(holidaysDict):
			if int(holidayDate) < day:
				holidaysDict.pop(holidayDate)

	jsonData["holidaysList"].update(holidaysDict)
	sendDataToJSON('log.json', jsonData)

# loads the time table from the excel sheet and then updates the classes today
def loadTimeTable():
	classTimeTableLocation = config.TIME_TABLE
	timetablewb = openpyxl.load_workbook(classTimeTableLocation) 
	sheet = timetablewb.active
	maxColumn = sheet.max_column
	maxRow = sheet.max_row
	completeTimeTable = {}
	for i in range(1, maxRow + 1):
		rowValues = []
		keyValue = sheet.cell(row = i, column = 1)
		for j in range(2, maxColumn + 1):
			cellValue = sheet.cell(row = i, column = j)
			if cellValue.value != None:
				rowValues.append(cellValue.value)
		completeTimeTable[keyValue.value] = rowValues
	jsonData = fetchDataFromJSON('log.json')
	jsonData["completeTimeTable"].update(completeTimeTable)
	sendDataToJSON('log.json', jsonData)
	classesToday(printHoliday = False)

def sendTimetable():
	loadTimeTable()
	log = fetchDataFromJSON('log.json')
	timetableData = log["completeTimeTable"]
	timings = timetableData['Timings']
	s = '__**Time Table**__\n'
	for key in timetableData:
		if key != 'Timings':
			datalist = timetableData[key]
			period = 0
			s += '__**' + key + '**__' + ':' + '\n'
			for i in range(len(datalist)):
				s += timings[period] + ' : ' + datalist[period] + '\n'
				period += 1
			s += '\n'
	return s

# returns the status of the classwork
# if no class is going on, then it returns False which means that classes are scheduled later
# if classwork is completed then it returns True which means all classes are already completed
# if classwork is going on, then it returns -1
def classStatus():
	jsonData = fetchDataFromJSON('log.json')
	todaysTimeTable = jsonData["todaysTimeTable"]
	timings = list(todaysTimeTable.keys())
	startTime = timings[0][:5]
	endTime = timings[-1][8:]
	h, m, s = str(datetime.now().time()).split(':')
	presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
	classstartTime = timedelta(hours = int(startTime[:2]), minutes = int(startTime[3:5]))
	classendTime = timedelta(hours = int(endTime[:2]), minutes = int(endTime[3:]))
	if presentTime < classstartTime:
		return False
	if presentTime > classstartTime and presentTime < classendTime:	
		return -1
	if not presentTime < classendTime:
		return True

# makes a schedule  for classes today and prints  
def classesToday(printHoliday = True):
	classes = []
	updateholidaysList()
	jsonData = fetchDataFromJSON('log.json')
	holidays = jsonData["holidaysList"]
	dateAndTime = datetime.now()
	day = dateAndTime.day
	weekDay = dateAndTime.today().strftime('%A')
	if str(day) in holidays and printHoliday:
		print('[' + str(datetime.now().strftime("%H:%M:%S")) + '] ' + "\n\tToday is a holiday due to " + holidays[str(day)])
	else:
		classes = []
		classtime = []
		classesToday = jsonData["completeTimeTable"][weekDay]
		timings = jsonData["completeTimeTable"]["Timings"]
		timings = timings[:len(classesToday)]
		prevClass = ''
		for i in range(len(classesToday)):
			if classesToday[i] == prevClass:
				classtime[-1] = classtime[-1][:8] + timings[i][8:]
			else :
				classes.append(classesToday[i])
				classtime.append(timings[i])
				prevClass = classesToday[i]
		t = {classtime[i]: classes[i] for i in range(len(classtime))}
		jsonData = fetchDataFromJSON('log.json')
		classLoginLog = jsonData["log"]["classStatus"]
		classesAlreadyAttended = list(classLoginLog.keys())
		jsonData["todaysTimeTable"] = t
		sendDataToJSON('log.json', jsonData)

# returns current class 
# if no class is going on or if the current class is already attended then it will return None
def whichClass(nextClass = False, nextClassTime = False, currentClassTime = False) :
	loadTimeTable()
	jsonData = fetchDataFromJSON('log.json')
	subjects = jsonData["todaysTimeTable"]
	classLoginLog = jsonData["log"]["classStatus"]
	classesAlreadyAttended = list(classLoginLog.keys())
	nextClassFlag = False
	h, m, s = str(datetime.now().time()).split(':')
	presentTime = timedelta(hours = int(h), minutes = int(m), seconds = int(float(s)))
	for i in subjects:
		startTime = timedelta(hours = int(i[0:2]), minutes = int(i[3:5]))
		endTime = timedelta(hours = int(i[8:10]), minutes = int(i[11:]))
		if nextClassFlag and (not nextClassTime):
			return subjects[i]
		if nextClassFlag and nextClassTime:
			return (i, subjects[i])
		if nextClass:
			if presentTime < endTime and presentTime > startTime:
				nextClassFlag = True
				continue
		elif  presentTime < endTime and presentTime > startTime and (not(i in classesAlreadyAttended)):	
			if currentClassTime:
				return (i, subjects[i])
			return subjects[i]	
	if nextClassTime or currentClassTime:
		return (None, None)
	return None

# reverts to the original timetable
def revertTimeTable():
	jsonData = fetchDataFromJSON('log.json')
	totalUpdates = len(list(jsonData["tempTimetableUpdate"].keys()))
	if totalUpdates > 0:
		for i in range(totalUpdates):
			day = jsonData["tempTimetableUpdate"][str(i)]["day"]
			period = jsonData["tempTimetableUpdate"][str(i)]["period"]
			classToUpdate = jsonData["tempTimetableUpdate"][str(i)]["previousClass"]
			updateTimeTable(day, period, classToUpdate)
		jsonData["tempTimetableUpdate"] = {}
		sendDataToJSON('log.json', jsonData)

# updates the time table by taking day, period and class to update
# day should be string of any format. eg: "Monday", "monday", "MONday"
# period should be passed as stored in excel sheet
# class that you want to replace with
# if you want to remove a class then just pass the class with ''
def updateTimeTable(day, period, classToUpdate = None, previousClass = False):
	classTimeTableLocation = config.TIME_TABLE
	timetablewb = openpyxl.load_workbook(classTimeTableLocation) 
	sheet = timetablewb.active
	day = day.lower()
	log = fetchDataFromJSON('log.json')
	i, val, asciiVal = 1, 2, 66
	rowValues, colValues, colDict = {}, {}, {}
	for key in list(log['completeTimeTable'].keys())[1:] :
		key = key.lower()
		rowValues[key] = str(val)
		colValues[key] = chr(asciiVal)
		colDict[str(i)] = key
		i += 1
		asciiVal += 1	
		val += 1	
	
	prevClass =  sheet[colValues[colDict[period]] + str(rowValues[day])].value
	sheet[colValues[colDict[period]] + str(rowValues[day])] = classToUpdate
	timetablewb.save(classTimeTableLocation)
	if previousClass:
		return prevClass
	
# prints log data
def printLog():
	logData = fetchDataFromJSON('log.json')
	return json.dumps(logData, indent = 4)

# returns the count of members joined before joining the class
def membersAlreadyJoinedCount(text):
	if text == 'No one else is here':
		return 0
	count = 0
	numberFromText = [int(i) for i in text.split() if i.isdigit()]
	commasCount = text.count(',')
	andCount = text.count('and')
	if len(numberFromText) > 0:
		count = numberFromText[0]
		andCount = 0
	count = count + commasCount + andCount + 1
	return count
	
# updates member count to log.json
def updateMembersCount(count):
	logData = fetchDataFromJSON('log.json')
	logData["log"]["membersCount"] = count
	sendDataToJSON('log.json', logData)
		
# checks whether google account is logged in or not
def checklogin(context):
	if os.path.exists("google.pkl"):
		cookies = pickle.load(open("google.pkl", "rb"))
		driver.get('https://apps.google.com/meet/')
		for cookie in cookies:
			driver.add_cookie(cookie)

		# loaded screen recorder 
		#driver.get('file:///home/koteshrv/Desktop/googlemeetbot/screenrecord.html')

		# Open a new window
		#driver.execute_script("window.open('');")
		
		# Switch to the new window and open new URL
		#driver.switch_to.window(driver.window_handles[1])

	else:
		context.bot.send_message(chat_id = config.TELEGRAM_USER_ID, text= "You're not logged in. Please run /login command to login. Then try again!")
		discordAndPrint("You're not logged in. Please run /login command in telegram to login. Then try again!")
		return

def takeScreenshot(context):
	fileName = "status@" + str(datetime.now().strftime("%H_%M_%S")) + ".png"
	driver.save_screenshot(fileName)
	context.bot.send_chat_action(chat_id = config.TELEGRAM_USER_ID, action = ChatAction.UPLOAD_DOCUMENT)
	context.bot.sendDocument(chat_id = config.TELEGRAM_USER_ID, document = open(fileName, 'rb'))
	os.remove(fileName)

def sendToTelegram(context, message):
	context.bot.send_chat_action(chat_id = config.TELEGRAM_USER_ID, action = ChatAction.TYPING)
	context.bot.send_message(chat_id = config.TELEGRAM_USER_ID, text = message)
	discordAndPrint('Sent a message successfully!')


